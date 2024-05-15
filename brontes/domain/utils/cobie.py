
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from io import BytesIO
from typing import Tuple, Dict, List
from rdflib import Literal, RDF, URIRef, Graph

from brontes.domain.models import COBieSpreadsheet, Type, Category, Floor, Space, Component, System, Facility
from brontes.utils import create_uri
from brontes.infrastructure.db.knowledge_graph import KnowledgeGraph 

def parse_spreadsheet(facility: Facility, file: str | bytes) -> COBieSpreadsheet:
  """
  Parse a COBie spreadsheet file into a COBieSpreadsheet object.

  Args:
  - file_content: The content of the COBie spreadsheet file.

  Returns:
  - A COBieSpreadsheet object
  """
  try:
    facility_uri = facility.uri
    df = pd.read_excel(BytesIO(file), engine='openpyxl', sheet_name=None)

    # Floors Sheet
    floors: List[Floor] = []
    for _, row in df['Floor'].iterrows():
      floor_uri = f"{facility_uri}/floor/{create_uri(row['Name'])}"
      floor = Floor(
        uri=floor_uri,
        name=row['Name'],
        description=row['Description'],
        elevation=row['Elevation'],
        height=row['Height']
      )
      floors.append(floor)

    # Spaces Sheet
    spaces: List[Space] = []
    for _, row in df['Space'].iterrows():
      space_uri = f"{facility_uri}/space/{create_uri(row['Name'])}"
      floor_uri = f"{facility_uri}/floor/{create_uri(row['FloorName'])}"
      category_uri = f"https://syyclops.com/categorySpace/{create_uri(row['Category'])}"
      category = Category(uri=category_uri, hasStringValue=row['Category'])
      space = Space(
        uri=space_uri,
        name=row['Name'],
        floor=next(floor for floor in floors if floor.uri == floor_uri),
        description=row['Description'],
        extIdentifier=row['ExtIdentifier'],
        grossArea=row['GrossArea'],
        netArea=row['NetArea'],
        category=category
      )
      spaces.append(space)

    # Types Sheet
    types: List[Type] = []
    for _, row in df['Type'].iterrows():
      type_uri = f"{facility_uri}/type/{create_uri(row['Name'])}"
      category_uri = f"https://syyclops.com/categoryProduct/{create_uri(row['Category'])}"
      category = Category(uri=category_uri, hasStringValue=row['Category'])
      cobie_type = Type(
        uri=type_uri, 
        name=row['Name'],
        description=row['Description'],
        modelNumber=row['ModelNumber'],
        extIdentifier=row['ExtIdentifier'],
        category=category
      )
      types.append(cobie_type)

    # Components Sheet
    components: List[Component] = []
    for _, row in df['Component'].iterrows():
      component_uri = f"{facility_uri}/component/{create_uri(row['Name'])}"
      type_uri = f"{facility_uri}/type/{create_uri(row['TypeName'])}"
      space_uri = f"{facility_uri}/space/{create_uri(row['Space'])}"
      component = Component(
        uri=component_uri,
        name=row['Name'],
        description=row['Description'],
        type=next(cobie_type for cobie_type in types if cobie_type.uri == type_uri),
        space=next((space for space in spaces if space.uri == space_uri), None),
        extIdentifier=row['ExtIdentifier']
      )
      components.append(component)

    # Systems Sheet
    systems: List[System] = []
    for _, row in df['System'].iterrows():
      system_uri = f"{facility_uri}/system/{create_uri(row['Name'])}"
      component_uri = f"{facility_uri}/component/{create_uri(row['ComponentNames'])}"

      component = next(component for component in components if component.uri == component_uri)

      # If system already exists, append the component to the system
      if system_uri in [system.uri for system in systems]:
        system = next(system for system in systems if system.uri == system_uri)
        system.components.append(component)
      else:
        system = System(
          uri=system_uri,
          name=row['Name'],
          components=[component],
          description=row['Description']
        )
        systems.append(system)

    return COBieSpreadsheet(floors=floors, spaces=spaces, types=types, components=components, systems=systems)
  except Exception as e:
    raise e
  
def upload_to_graph(g: Graph, spreadsheet: COBieSpreadsheet) -> str:
  """
  Upload a COBie spreadsheet to a RDF graph store.

  Args:
  - g: A RDF graph store
  - spreadsheet: A COBieSpreadsheet object

  Returns:
  - An RDF string
  """

  # Define common namespaces
  COBIE = KnowledgeGraph.prefixes["cobie"]
  A = RDF.type

  try:
    for floor in spreadsheet.floors:
      floor_node = URIRef(floor.uri)
      g.add((floor_node, A, COBIE.Floor))
      g.add((floor_node, COBIE.name, Literal(floor.name)))
      g.add((floor_node, COBIE.description, Literal(floor.description)))
      g.add((floor_node, COBIE.elevation, Literal(floor.elevation)))
      g.add((floor_node, COBIE.height, Literal(floor.height)))

    for space in spreadsheet.spaces:
      space_node = URIRef(space.uri)
      # Create properties for the space
      g.add((space_node, A, COBIE.Space))
      g.add((space_node, COBIE.name, Literal(space.name)))
      g.add((space_node, COBIE.description, Literal(space.description)))
      g.add((space_node, COBIE.extIdentifier, Literal(space.extIdentifier)))
      g.add((space_node, COBIE.grossArea, Literal(space.grossArea)))
      g.add((space_node, COBIE.netArea, Literal(space.netArea)))
      # Create relationships
      g.add((space_node, COBIE.category, URIRef(space.category.uri)))

    for cobie_type in spreadsheet.types:
      type_node = URIRef(cobie_type.uri)
      # Create properties for the type
      g.add((type_node, A, COBIE.Type))
      g.add((type_node, COBIE.name, Literal(cobie_type.name)))
      g.add((type_node, COBIE.description, Literal(cobie_type.description)))
      g.add((type_node, COBIE.modelNumber, Literal(cobie_type.modelNumber)))
      g.add((type_node, COBIE.extIdentifier, Literal(cobie_type.extIdentifier)))
      # Create relationships
      g.add((type_node, COBIE.category, URIRef(cobie_type.category.uri)))

    for component in spreadsheet.components:
      component_node = URIRef(component.uri)
      # Create properties for the component
      g.add((component_node, A, COBIE.Component))
      g.add((component_node, COBIE.name, Literal(component.name)))
      g.add((component_node, COBIE.description, Literal(component.description)))
      g.add((component_node, COBIE.extIdentifier, Literal(component.extIdentifier)))
      g.add((component_node, COBIE.serialNumber, Literal(component.serialNumber)))
      # Create relationships
      g.add((component_node, COBIE.type, URIRef(component.type.uri)))
      if component.space:
        g.add((component_node, COBIE.space, URIRef(component.space.uri)))

    for system in spreadsheet.systems:
      system_node = URIRef(system.uri)
      # Create properties for the system
      g.add((system_node, A, COBIE.System))
      g.add((system_node, COBIE.name, Literal(system.name)))
      g.add((system_node, COBIE.description, Literal(system.description)))
      # Create relationships
      for component in system.components:
        g.add((system_node, COBIE.componentNames, URIRef(component.uri)))
  except Exception as e:
    raise e

def validate_spreadsheet(file_content: bytes) -> Tuple[bool, Dict, bytes]:
  """
  Validate a COBie spreadsheet. Refer to COBie_validation.pdf in docs/ for more information.

  Returns: 
  - errors_found: A boolean indicating whether or not errors were found in the spreadsheet.
  - errors: A dictionary containing all the errors found in the spreadsheet.
  - updated_file: The file path of the updated spreadsheet with the errors highlighted in red.
  """
  errors = {
    "Expected sheet not found in spreadsheet.": [],
    "More than one record found in Facility sheet.": [],
    "Empty or N/A cells found in column A of sheet.": [],
    "Duplicate names found in column A of sheet.": [],
    "Space is not linked to a value in the first column of the Floor tab.": [],
    "Not every Type record has a category.": [],
    "Component is not linked to an existing Type.": [],
    "Component is not linked to an existing Space.": []
  }

  errors_found = False

  # Open COBie spreadsheet
  df = pd.read_excel(BytesIO(file_content), engine='openpyxl', sheet_name=None) 
  wb = openpyxl.load_workbook(BytesIO(file_content))

  expected_sheets = ['Facility', 'Floor', 'Space', 'Type', 'Component', 'Attribute', 'System']
  # Check to make sure the spreadsheet has the correct sheets     
  for sheet in expected_sheets:
    if sheet not in df.keys():
      errors["Expected sheet not found in spreadsheet."].append({
          "sheet": sheet,
      })
      errors_found = True
  if errors_found: return errors_found, errors, file_content

  # Make sure there is only one record in the Facility sheet
  if len(df['Facility']) > 1:
    errors["More than one record found in Facility sheet."].append({
      "sheet": "Facility",
      "row": 1,
      "column": 1
    })
    errors_found = True

    # highlight the sheet in red
    cell = wb['Facility'].cell(row=1, column=1)
    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type = "solid")

  # No empty or N/A cells are present in column A of any sheet
  for sheet in expected_sheets:
    for idx, value in enumerate(df[sheet]['Name']):
      if pd.isnull(value) or value == None or value == "N/A" or pd.isna(value):
        errors["Empty or N/A cells found in column A of sheet."].append({
          "sheet": sheet,
          "row": idx + 2,
          "column": 1
        })
        errors_found = True

        # highlight the cell in red
        cell = wb[sheet].cell(row=idx + 2, column=1)
        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type = "solid")

  # Check Floor, Space, Type, Component sheets for duplicate names in column A
  for sheet in ['Floor', 'Space', 'Type', 'Component']:
    seen_names = set()
    for idx, name in enumerate(df[sheet]['Name']):
      if name in seen_names:
        errors["Duplicate names found in column A of sheet."].append({
          "sheet": sheet,
          "row": idx + 2,
          "column": 1
        })  
        errors_found = True

        cell = wb[sheet].cell(row=idx + 2, column=1)
        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type = "solid")
      else:
        seen_names.add(name)

  # Space Tab
  # Every value is linked to a value in the first column of the Floor tab
  for idx, space in enumerate(df['Space']['FloorName']):
    if space not in df['Floor']['Name'].values:
      errors["Space is not linked to a value in the first column of the Floor tab."].append({
        "sheet": "Space",
        "row": idx + 2,
        "column": 5
      })  
      errors_found = True

      cell = wb['Space'].cell(row=idx + 2, column=5)
      cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type = "solid")

  # Type Tab
  # Every record has a category
  for idx, type in enumerate(df['Type']['Category']):
    if pd.isnull(type):
      errors["Not every Type record has a category."].append({
        "sheet": "Type",
        "row": idx + 2,
        "column": 4
      })
      errors_found = True

      cell = wb['Type'].cell(row=idx + 2, column=4)
      cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type = "solid")
      
  # Component Tab
  # Every record is linked to a existing Type
  for idx, component in enumerate(df['Component']['TypeName']):
    if component not in df['Type']['Name'].values:
      errors["Component is not linked to an existing Type."].append({
        "sheet": "Component",
        "row": idx + 2,
        "column": 4
      })
      errors_found = True

      cell = wb['Component'].cell(row=idx + 2, column=4)
      cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type = "solid")

  # Every record is linked a to existing Space
  for idx, space_name in enumerate(df['Component']['Space']):
    # Check if cell is valid
    if pd.isnull(space_name):
      errors["Component is not linked to an existing Space."].append({
        "sheet": "Component",
        "row": idx + 2,
        "column": 5
      })
      errors_found = True

      cell = wb['Component'].cell(row=idx + 2, column=5)
      cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type = "solid")
      continue
    # Split by "," to get all spaces and remove whitespace
    spaces = [space.strip() for space in space_name.split(",")]
    for space in spaces:
      if space not in df['Space']['Name'].values:
        errors["Component is not linked to an existing Space."].append({
          "sheet": "Component",
          "row": idx + 2,
          "column": 5
        })
        errors_found = True

        cell = wb['Component'].cell(row=idx + 2, column=5)
        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type = "solid")

  # Remove all the empty lists from the errors dict
  errors = {key: value for key, value in errors.items() if value}

  # Save the workbook
  content = BytesIO()
  wb.save(content)
  content.seek(0)

  return errors_found, errors, content.getvalue()